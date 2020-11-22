# py -m pip install -r requirements.txt 
from openpyxl import Workbook, load_workbook
import datetime
import os

FILENAME = "Daten.xlsx"
DATEFORMAT = "%Y-%m-%d %X"

def clear(): os.system('cls') #on Windows System
clear()

try:
    wb = load_workbook(FILENAME) # loads existing Workbook
    ws = wb.active # grab the active worksheet
except Exception:
    wb = Workbook() # creates new Workbook
    ws = wb.active  # grab the active worksheet
    ws.append(["Zeit", "Name des Käufers","Name des Empängers","Jahrgang des Empängers","Klasse des Empängers","Vollmilch Menge", "Zartbitter Menge", "Anonym"])


def Eingabe():

    nameBuyer = input("\nName des Käufers: ")

    retry = False
    while retry == False:
        retry = True
        try:
            entrys = int(input("\nAnzahl der Käufe: "))
        except:
            print("\nWARNING: Anzahl der Käufe muss eine -Zahl- sein")
            retry = False



    for x in range (0, entrys):
        
        nameReciever = input("\nName des Empängers "+ str(x+1) +": ")

        jahrgang = input("Jahrgang des Empängers "+ str(x+1) +": ")
        grade = input("Klasse des Empängers "+ str(x+1) +": ")

        retryproducts = False

        while retryproducts == False:
            retryproducts = True
            try:
                vollmilch = int(input("Vollmilch groß Menge: "))
                zartbitter = int(input("Zartbitter klein Menge: "))
            except:
                print("\nProduktmengen müssen eine -Zahl- sein")
                retryproducts = False


        anonym = input("Anonym: (J/N): ").lower().startswith("j")

        row = [datetime.datetime.now().strftime(DATEFORMAT), nameBuyer, nameReciever, jahrgang, grade, vollmilch, zartbitter, anonym]

        if input("\nZum Speichern ENTER drücken, zum Abbrechen 'STRG + C': ").lower().startswith("n") == False:
            ws.append(row)
            wb.save(FILENAME)
            if x == (entrys-1):
                clear()
                print("\nSaved")
            else:
                print("\nNächster Eintrag '"+ str(x+2) +"'")
        else:
            clear()
            print("\nCanceled")
            x = x + 1


if __name__ == "__main__":
    while True:
        try: 
            Eingabe()
        except KeyboardInterrupt:
            clear()
            print("\nCanceled")